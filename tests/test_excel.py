import os
import tempfile
import openpyxl
import polars as pl
from app.core.excel_reader import read_excel_file, get_excel_sheets

def test_excel_reading():
    # Generate temporary Excel file
    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "Clientes"
    ws1.append(["nome", "cpf", "cep"])
    # Write cpf as numeric, it should be converted to string without .0
    ws1.append(["João", 1234567890, 1001000.0])
    
    ws2 = wb.create_sheet("Vazia")
    
    ws3 = wb.create_sheet("Vendas")
    ws3.append(["codigo", "valor"])
    ws3.append(["C001", 10.5])
    
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        temp_name = f.name
    
    try:
        wb.save(temp_name)
        wb.close()
        
        # Test sheets listing
        sheets = get_excel_sheets(temp_name)
        assert sheets == ["Clientes", "Vazia", "Vendas"]
        
        # Test read first sheet by default (ignores Vazia since Clientes is first)
        dfs_single = read_excel_file(temp_name, process_all_sheets=False, preserve_sensitive=True)
        assert list(dfs_single.keys()) == ["Clientes"]
        df_cli = dfs_single["Clientes"]
        assert df_cli.height == 1
        assert df_cli.schema["cpf"] == pl.String
        assert df_cli.schema["cep"] == pl.String
        # Verify zeros / values are kept as integer strings without decimals
        assert df_cli["cpf"][0] == "1234567890"
        assert df_cli["cep"][0] == "1001000"
        
        # Test read all sheets (Vazia should be ignored as it has no columns/rows)
        dfs_all = read_excel_file(temp_name, process_all_sheets=True, preserve_sensitive=True)
        assert "Clientes" in dfs_all
        assert "Vendas" in dfs_all
        assert "Vazia" not in dfs_all
        assert dfs_all["Vendas"].schema["codigo"] == pl.String  # sensitive
        assert dfs_all["Vendas"]["codigo"][0] == "C001"
        assert dfs_all["Vendas"]["valor"][0] == 10.5  # float is kept float because valor is not sensitive
        
    finally:
        if os.path.exists(temp_name):
            os.remove(temp_name)
network_access_required = False
